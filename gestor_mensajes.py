import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import time
import os
from openpyxl import load_workbook

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service

# ─────────────────────────────
# CONFIG
# ─────────────────────────────

BRAVE_PATHS = [
    r"C:/Program Files/BraveSoftware/Brave-Browser/Application/brave.exe",
    r"C:/Program Files (x86)/BraveSoftware/Brave-Browser/Application/brave.exe",
    os.path.expanduser(r"~\AppData\Local\BraveSoftware\Brave-Browser\Application\brave.exe"),
]

PROFILE_DIR = os.path.expanduser("~/.brave_selenium_profile")
DRIVER_PATH = os.path.expanduser("~/.wdm/brave_driver/chromedriver.exe")

# ─────────────────────────────
# DRIVER — se engancha a la sesión existente o abre una vez
# ─────────────────────────────

_DRIVER = None

def get_driver(log):
    global _DRIVER
    if _DRIVER:
        try:
            _ = _DRIVER.current_url   # verifica que sigue vivo
            return _DRIVER
        except Exception:
            _DRIVER = None

    brave = next((p for p in BRAVE_PATHS if os.path.exists(p)), None)
    if not brave:
        raise Exception("Brave no encontrado. Verifica la ruta en CONFIG.")

    options = webdriver.ChromeOptions()
    options.binary_location = brave
    # Reutiliza el perfil donde ya tienes WhatsApp Web abierto
    options.add_argument(f"--user-data-dir={PROFILE_DIR}")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    # Evita que abra ventana de restauración de sesión
    options.add_argument("--no-first-run")
    options.add_argument("--no-default-browser-check")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)

    _DRIVER = webdriver.Chrome(service=Service(DRIVER_PATH), options=options)
    log("🌐 Navegador iniciado con tu sesión guardada")
    return _DRIVER

# ─────────────────────────────
# ENVÍO — navega dentro de la misma pestaña
# ─────────────────────────────

def detectar_error_chat(driver):
    errores = [
        "Phone number shared via url is invalid",
        "Número de teléfono compartido mediante la URL no es válido",
        "This account cannot use WhatsApp",
        "invalid phone",
    ]
    body = driver.page_source
    return any(e.lower() in body.lower() for e in errores)


def enviar_mensajes(driver, wait, numero, mensaje, veces=1):
    """Envía 'mensaje' a 'numero' exactamente 'veces' veces."""
    url = f"https://web.whatsapp.com/send?phone={numero}&app_absent=0"
    driver.get(url)   # navega en la misma pestaña, sin abrir nuevas

    # Espera el cuadro de texto
    caja = wait.until(EC.presence_of_element_located(
        (By.XPATH, '//div[@title="Type a message" or @title="Escribe un mensaje"]')
    ))

    if detectar_error_chat(driver):
        raise ValueError("Número inválido o bloqueado")

    caja.click()
    for _ in range(veces):
        caja.send_keys(mensaje)
        caja.send_keys(Keys.ENTER)
        time.sleep(0.4)   # pequeña pausa entre mensajes repetidos

# ─────────────────────────────
# EXCEL
# ─────────────────────────────

def cargar_numeros_excel(ruta):
    wb = load_workbook(ruta)
    ws = wb.active
    nums = []
    for cell in ws["A"]:
        if cell.value:
            nums.append(str(cell.value).strip())
    return nums

# ─────────────────────────────
# APP UI
# ─────────────────────────────

VERDE   = "#25D366"
VERDE2  = "#128C7E"
BG      = "#0d0d0d"
BG2     = "#161616"
BG3     = "#1e1e1e"
TEXT    = "#e8e8e8"
MUTED   = "#666"
RED     = "#e05555"

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("WhatsApp Bulk Sender")
        self.root.geometry("560x800")
        self.root.minsize(520, 600)
        self.root.configure(bg=BG)
        self.root.resizable(True, True)

        self.numeros = []
        self.enviando = False

        self._build_ui()

    # ── Construcción de la interfaz ──────────────────────────────────

    def _build_ui(self):
        # Título
        hdr = tk.Frame(self.root, bg=BG)
        hdr.pack(fill="x", pady=(22, 4))
        tk.Label(hdr, text="📨", font=("Segoe UI Emoji", 28),
                 bg=BG).pack()
        tk.Label(hdr, text="WhatsApp Bulk Sender",
                 font=("Segoe UI", 18, "bold"),
                 bg=BG, fg=VERDE).pack()
        tk.Label(hdr, text="Envía mensajes masivos sin escanear QR",
                 font=("Segoe UI", 9), bg=BG, fg=MUTED).pack()

        # Separador
        tk.Frame(self.root, height=1, bg=BG3).pack(fill="x", pady=14)

        # Contenedor principal
        # Contenedor con scroll para que el botón nunca quede oculto
        canvas = tk.Canvas(self.root, bg=BG, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        body = tk.Frame(canvas, bg=BG, padx=30)
        body_win = canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_canvas_resize(e):
            canvas.itemconfig(body_win, width=e.width)
        canvas.bind("<Configure>", _on_canvas_resize)

        body.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))

        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(
            int(-1 * (e.delta / 120)), "units"))

        # ── Sección números ─────────────────────────────────────────
        self._section(body, "1. DESTINATARIOS")

        tabs = ttk.Notebook(body)
        tabs.pack(fill="x", pady=(0, 10))

        # Tab Excel
        tab_excel = tk.Frame(tabs, bg=BG2, padx=10, pady=8)
        tabs.add(tab_excel, text="  📂 Desde Excel  ")

        btn_cargar = self._btn(tab_excel, "Seleccionar archivo .xlsx",
                               self._cargar_excel)
        btn_cargar.pack(fill="x")
        self.lbl_excel = tk.Label(tab_excel, text="Ningún archivo cargado",
                                  bg=BG2, fg=MUTED, font=("Segoe UI", 9))
        self.lbl_excel.pack(anchor="w", pady=(4, 0))

        # Tab manual
        tab_manual = tk.Frame(tabs, bg=BG2, padx=10, pady=8)
        tabs.add(tab_manual, text="  ✏️ Pegar números  ")
        tk.Label(tab_manual, text="Un número por línea (ej: +51987654321)",
                 bg=BG2, fg=MUTED, font=("Segoe UI", 8)).pack(anchor="w")
        self.txt_nums = tk.Text(tab_manual, height=5,
                                bg=BG3, fg=TEXT,
                                insertbackground=VERDE,
                                font=("Consolas", 10),
                                relief="flat", padx=6, pady=6)
        self.txt_nums.pack(fill="x", pady=(4, 0))

        self.tabs = tabs

        # ── Mensaje ─────────────────────────────────────────────────
        self._section(body, "2. MENSAJE")
        self.txt_msg = tk.Text(body, height=5,
                               bg=BG2, fg=TEXT,
                               insertbackground=VERDE,
                               font=("Segoe UI", 10),
                               relief="flat", padx=8, pady=8)
        self.txt_msg.pack(fill="x")
        self.lbl_chars = tk.Label(body, text="0 caracteres",
                                  bg=BG, fg=MUTED, font=("Segoe UI", 8))
        self.lbl_chars.pack(anchor="e", pady=(2, 0))
        self.txt_msg.bind("<KeyRelease>", self._actualizar_chars)

        # ── Opciones ────────────────────────────────────────────────
        self._section(body, "3. OPCIONES")
        opts = tk.Frame(body, bg=BG)
        opts.pack(fill="x", pady=(0, 10))
        opts.columnconfigure(0, weight=1)
        opts.columnconfigure(1, weight=1)

        # Repeticiones
        tk.Label(opts, text="Nº de veces por contacto",
                 bg=BG, fg=TEXT, font=("Segoe UI", 9)).grid(
                 row=0, column=0, sticky="w")
        spin_frame = tk.Frame(opts, bg=BG)
        spin_frame.grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.spin_veces = tk.Spinbox(
            spin_frame, from_=1, to=20, width=5,
            bg=BG2, fg=VERDE, buttonbackground=BG3,
            relief="flat", font=("Segoe UI", 11, "bold"),
            insertbackground=VERDE)
        self.spin_veces.pack()

        # Delay
        tk.Label(opts, text="Delay entre contactos (seg)",
                 bg=BG, fg=TEXT, font=("Segoe UI", 9)).grid(
                 row=0, column=1, sticky="w", padx=(20, 0))
        self.scale_delay = tk.Scale(
            opts, from_=0, to=10, orient="horizontal",
            bg=BG, fg=VERDE, highlightthickness=0,
            troughcolor=BG3, activebackground=VERDE2,
            length=160)
        self.scale_delay.set(2)
        self.scale_delay.grid(row=1, column=1, sticky="w", padx=(20, 0))

        # ── Progreso ─────────────────────────────────────────────────
        tk.Frame(body, height=1, bg=BG3).pack(fill="x", pady=10)

        self.lbl_estado = tk.Label(body, text="Listo para enviar",
                                   bg=BG, fg=VERDE,
                                   font=("Segoe UI", 10, "bold"))
        self.lbl_estado.pack(anchor="w")

        self.progress = ttk.Progressbar(body, length=480,
                                        style="green.Horizontal.TProgressbar")
        self.progress.pack(fill="x", pady=(6, 10))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("green.Horizontal.TProgressbar",
                        troughcolor=BG3, background=VERDE,
                        bordercolor=BG, lightcolor=VERDE, darkcolor=VERDE2)

        # ── Botón enviar ─────────────────────────────────────────────
        self.btn_enviar = self._btn(body, "🚀  Iniciar Envío",
                                    self._iniciar, big=True)
        self.btn_enviar.pack(fill="x", pady=(0, 8))

        # ── Log ──────────────────────────────────────────────────────
        self.log_box = tk.Text(body, height=7,
                               bg=BG2, fg=VERDE,
                               font=("Consolas", 9),
                               relief="flat", padx=6, pady=6,
                               state="disabled")
        self.log_box.pack(fill="both", pady=(0, 16))

    # ── Helpers UI ───────────────────────────────────────────────────

    def _section(self, parent, title):
        tk.Label(parent, text=title,
                 font=("Segoe UI", 8, "bold"),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(12, 4))

    def _btn(self, parent, texto, cmd, big=False):
        size = 11 if big else 10
        pad  = 12 if big else 8
        return tk.Button(parent, text=texto, command=cmd,
                         bg=VERDE2, fg="white", activebackground=VERDE,
                         activeforeground="white",
                         font=("Segoe UI", size, "bold"),
                         relief="flat", cursor="hand2",
                         pady=pad)

    def _actualizar_chars(self, event=None):
        n = len(self.txt_msg.get("1.0", "end").strip())
        self.lbl_chars.config(text=f"{n} caracteres")

    def _log(self, texto):
        self.log_box.config(state="normal")
        self.log_box.insert("end", texto + "\n")
        self.log_box.see("end")
        self.log_box.config(state="disabled")

    def _set_estado(self, texto, color=VERDE):
        self.lbl_estado.config(text=texto, fg=color)

    # ── Cargar Excel ─────────────────────────────────────────────────

    def _cargar_excel(self):
        ruta = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not ruta:
            return
        self.numeros = cargar_numeros_excel(ruta)
        nombre = os.path.basename(ruta)
        self.lbl_excel.config(
            text=f"✅ {nombre}  —  {len(self.numeros)} números",
            fg=VERDE)
        self._log(f"📂 Excel cargado: {len(self.numeros)} contactos")

    # ── Obtener números según pestaña activa ─────────────────────────

    def _obtener_numeros(self):
        tab = self.tabs.index(self.tabs.select())
        if tab == 0:
            return list(self.numeros)
        else:
            raw = self.txt_nums.get("1.0", "end").strip()
            return [l.strip() for l in raw.splitlines() if l.strip()]

    # ── Iniciar envío ────────────────────────────────────────────────

    def _iniciar(self):
        if self.enviando:
            messagebox.showinfo("En proceso", "Ya hay un envío en curso.")
            return

        numeros = self._obtener_numeros()
        if not numeros:
            messagebox.showwarning("Sin destinatarios",
                                   "Carga un Excel o pega números en la pestaña manual.")
            return

        msg = self.txt_msg.get("1.0", "end").strip()
        if not msg:
            messagebox.showwarning("Sin mensaje", "Escribe el mensaje a enviar.")
            return

        try:
            veces = int(self.spin_veces.get())
        except ValueError:
            veces = 1

        threading.Thread(
            target=self._enviar_hilo,
            args=(numeros, msg, veces),
            daemon=True
        ).start()

    # ── Hilo de envío ─────────────────────────────────────────────────

    def _enviar_hilo(self, numeros, mensaje, veces):
        self.enviando = True
        self.btn_enviar.config(state="disabled", bg="#555")
        self._set_estado("Iniciando navegador…")

        try:
            driver = get_driver(self._log)
        except Exception as e:
            messagebox.showerror("Error de navegador", str(e))
            self._set_estado("Error al iniciar", RED)
            self.enviando = False
            self.btn_enviar.config(state="normal", bg=VERDE2)
            return

        wait = WebDriverWait(driver, 20)
        total = len(numeros)
        self.progress["maximum"] = total
        enviados = 0
        errores  = 0

        for i, num in enumerate(numeros, 1):
            if not num.startswith("+"):
                num = "+" + num

            self._set_estado(f"Enviando {i}/{total}  —  {num}")
            self.progress["value"] = i
            self.root.update_idletasks()

            try:
                enviar_mensajes(driver, wait, num, mensaje, veces)
                self._log(f"✅  {num}  ({veces}x)")
                enviados += 1
            except ValueError:
                self._log(f"⛔  {num}  — número inválido o bloqueado")
                errores += 1
            except Exception as e:
                self._log(f"❌  {num}  — {e}")
                errores += 1

            delay = self.scale_delay.get()
            if delay > 0:
                time.sleep(delay)

        resumen = f"✔ Finalizado: {enviados} enviados, {errores} errores"
        self._set_estado(resumen, VERDE if errores == 0 else "#f0a500")
        self._log(f"\n{resumen}\n{'─'*40}")

        self.enviando = False
        self.btn_enviar.config(state="normal", bg=VERDE2)


# ─────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()